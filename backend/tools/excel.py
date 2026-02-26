from typing import Any

from openpyxl import load_workbook


def parse_excel(file_path: str) -> list[dict[str, Any]]:
    """
    Parse an Excel (.xlsx) file and return rows as a list of dicts.

    The first row is used as column headers. All subsequent rows are
    returned as dicts keyed by those headers.

    Args:
        file_path: Path to the .xlsx file.

    Returns:
        List of row dicts.
    """
    wb = load_workbook(filename=file_path, read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        wb.close()
        return []

    # First row = headers, normalize to lowercase stripped strings
    headers = [str(h).strip().lower() if h else f"col_{i}" for i, h in enumerate(rows[0])]

    data = []
    for row in rows[1:]:
        row_dict = {}
        for i, cell in enumerate(row):
            key = headers[i] if i < len(headers) else f"col_{i}"
            row_dict[key] = cell
        data.append(row_dict)

    wb.close()
    return data
